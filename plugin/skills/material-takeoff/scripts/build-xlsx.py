#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
#   "pillow>=10.0",
#   "pydantic>=2.0",
# ]
# ///
"""
Render a CompleteModel JSON (from `mto build-model`) into the project's
standard Excel take-off workbook.

Usage:
    build-xlsx.py <complete.json> <out.xlsx>

Three sheets:
  * Warnings          — any issues from extraction + build-model (only if any)
  * Materials         — one block per MaterialGroup
  * Building Elements — one block per element-material pair

Layout (Materials / Building Elements) matches
nio-energy-material-take-off-oekobaudat.xlsx:
  rows 1-9   : project header block (org, project name, branding)
  row 10     : section banner
  row 11     : column headers
  row 12+    : title row per group/element, then N quantity rows
"""

from __future__ import annotations

import sys
from pathlib import Path

from models import (
    CompleteModel,
    Element,
    Material,
    MaterialGroup,
    MaterialMatch,
    Project,
    Quantities,
    Warning,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="118A96")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
ERR_FILL = PatternFill("solid", fgColor="F8D7DA")
WHITE_BOLD = Font(color="FFFFFF", bold=True, size=13)
TITLE_FONT = Font(size=16)
LINK_FONT = Font(color="0563C1", underline="single", size=13)
CENTER = Alignment(horizontal="center", vertical="center")
MIDDLE = Alignment(vertical="center")
THIN_WHITE = Side(border_style="thin", color="FFFFFF")
HEADER_BORDER = Border(
    top=THIN_WHITE, bottom=THIN_WHITE, left=THIN_WHITE, right=THIN_WHITE
)

NUMBER_FORMAT = "#,##0.00"

MATERIAL_HEADERS = [
    "Name In Model",
    "Ökobaudat Material Name",
    "Category",
    "Quantity Name",
    "Value",
    "Unit",
    "Weight (kg)",
    "Embodied Carbon (kg CO2e)",
    "Whole Life Carbon Estimate (kg CO2e)",
    "Conformity",
    "Url",
]

ELEMENT_HEADERS = [
    "Guid",
    "Name",
    "Type",
    "Material Name In Model",
    "Parent Name",
    "Ökobaudat Material Name",
    "Category",
    "Quantity Name",
    "Value",
    "Unit",
    "Weight (kg)",
    "Embodied Carbon (kg CO2e)",
    "Whole Life Carbon Estimate (kg CO2e)",
]

# Where the quantity-row block starts (1-based column index) per sheet.
QUANTITY_START_COL = {"Materials": 4, "Building Elements": 8}

# Columns (per sheet) that hold the URL — for hyperlink styling.
URL_COL = {"Materials": 11}  # Building Elements doesn't have a URL column.


# Columns that hold numeric quantities/LCA values (for number formatting).
# col, col+1=value, col+3=weight, col+4=EC, col+5=WLC (col is QUANTITY_START_COL).
def _numeric_cols(sheet_title: str) -> list[int]:
    base = QUANTITY_START_COL[sheet_title]
    return [base + 1, base + 3, base + 4, base + 5]


def main() -> int:
    if len(sys.argv) < 3:
        print("usage: build-xlsx.py <complete.json> <out.xlsx>", file=sys.stderr)
        return 2

    src, dst = Path(sys.argv[1]), Path(sys.argv[2])
    model = CompleteModel.model_validate_json(src.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("Summary")
    _project_header(summary_ws, model.project)
    _white_header_band(summary_ws, 6)
    _write_summary(summary_ws, model)

    materials_ws = wb.create_sheet("Materials")
    _project_header(materials_ws, model.project)
    _white_header_band(materials_ws, len(MATERIAL_HEADERS))
    _section_banner(materials_ws, "MATERIALS", MATERIAL_HEADERS)
    _write_material_groups(materials_ws, model.material_groups)
    _finalize(materials_ws)

    elements_ws = wb.create_sheet("Building Elements")
    _project_header(elements_ws, model.project)
    _white_header_band(elements_ws, len(ELEMENT_HEADERS))
    _section_banner(elements_ws, "BUILDING ELEMENTS", ELEMENT_HEADERS)
    _write_elements(elements_ws, model.elements)
    _finalize(elements_ws)

    if model.match_decisions:
        matches_ws = wb.create_sheet("Matches")
        _write_matches(matches_ws, model.match_decisions)

    # Warnings last — visible but not the first thing the user sees.
    if model.warnings:
        warnings_ws = wb.create_sheet("Warnings")
        _write_warnings(warnings_ws, model.warnings)

    wb.save(dst)
    print(f"wrote {dst}", file=sys.stderr)
    return 0


# --- Warnings sheet ---------------------------------------------------------


def _write_warnings(ws: Worksheet, warnings: list[Warning]) -> None:
    """First sheet when anything went wrong — tells the user what's missing
    from the workbook so they don't trust the numbers blindly."""
    ws.cell(1, 1, "Warnings & notes").font = Font(bold=True, size=18)
    ws.cell(
        2,
        1,
        "These issues affected the take-off. Numbers in the other sheets exclude the affected elements / materials.",
    ).font = Font(size=12, italic=True)

    headers = ["Severity", "Code", "Message", "Details"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(4, i, h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = HEADER_BORDER

    for r, w in enumerate(warnings, start=5):
        fill = (
            ERR_FILL
            if w.level == "error"
            else WARN_FILL if w.level == "warning" else None
        )
        ws.cell(r, 1, w.level)
        ws.cell(r, 2, w.code)
        ws.cell(r, 3, w.message)
        ws.cell(r, 4, _format_context(w.context))
        if fill:
            for c in range(1, 5):
                ws.cell(r, c).fill = fill
        ws.row_dimensions[r].height = 25

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 60


# --- Summary sheet ----------------------------------------------------------


def _write_summary(ws: Worksheet, model: CompleteModel) -> None:
    """Executive view: project totals + top materials by embodied carbon.
    Sits after the project header block at row 10+."""
    # Totals: sum embodied/whole-life carbon across all groups' aggregated LCA.
    total_ec = 0.0
    total_wlc = 0.0
    rows: list[tuple[str, float, float]] = []
    for g in model.material_groups:
        if not g.aggregated_lca:
            continue
        # Pick a canonical quantity per group — prefer NetVolume, else first.
        net = next(
            (
                entry
                for entry in g.aggregated_lca
                if entry.reference_quantity == "NetVolume"
            ),
            None,
        )
        chosen = net or g.aggregated_lca[0]
        ec = chosen.embodied_carbon or 0.0
        wlc = chosen.whole_life_carbon or 0.0
        rows.append((g.name, ec, wlc))
        total_ec += ec
        total_wlc += wlc
    rows.sort(key=lambda r: r[1], reverse=True)

    # Banner spans cols B-F (under the logo + alongside the project info column)
    ws.cell(10, 2, "SUMMARY")
    for col in range(2, 7):
        c = ws.cell(10, col)
        c.fill = HEADER_FILL
        c.font = WHITE_BOLD
        c.alignment = CENTER
    ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=6)

    ws.cell(12, 2, "Total embodied carbon (kg CO2e)").font = Font(bold=True, size=13)
    _num(ws.cell(12, 4, total_ec))
    ws.cell(13, 2, "Total whole-life carbon (kg CO2e)").font = Font(bold=True, size=13)
    _num(ws.cell(13, 4, total_wlc))
    ws.cell(14, 2, "Material groups").font = Font(bold=True, size=13)
    ws.cell(14, 4, len(model.material_groups))
    ws.cell(15, 2, "Elements (with materials)").font = Font(bold=True, size=13)
    ws.cell(15, 4, sum(1 for e in model.elements if e.materials))

    # Top-N by embodied carbon
    ws.cell(18, 2, "Top materials by embodied carbon").font = Font(bold=True, size=14)
    for i, header in enumerate(
        [
            "Material",
            "Embodied Carbon (kg CO2e)",
            "Whole Life Carbon (kg CO2e)",
            "Share %",
        ],
        start=2,
    ):
        cell = ws.cell(19, i, header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER

    for i, (name, ec, wlc) in enumerate(rows[:10], start=20):
        ws.cell(i, 2, name)
        _num(ws.cell(i, 3, ec))
        _num(ws.cell(i, 4, wlc))
        if total_ec > 0:
            share = ws.cell(i, 5, ec / total_ec)
            share.number_format = "0.0%"

    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 12


# --- Matches sheet ----------------------------------------------------------

_CONFIDENCE_FILL = {
    "high": PatternFill("solid", fgColor="D4EDDA"),
    "medium": PatternFill("solid", fgColor="FFF3CD"),
    "low": PatternFill("solid", fgColor="F8D7DA"),
    "none": PatternFill("solid", fgColor="E9ECEF"),
}


def _write_matches(ws: Worksheet, matches: list[MaterialMatch]) -> None:
    """Audit trail: every distinct material → which DB record was chosen,
    confidence, and rationale. Lets the user spot-check the LLM's calls."""
    ws.cell(1, 1, "Match decisions").font = Font(bold=True, size=18)
    ws.cell(
        2,
        1,
        "Per-material match against the environmental database. Override here is informational only — re-run with corrected matches.json to regenerate.",
    ).font = Font(size=12, italic=True)

    headers = [
        "IFC Material Name",
        "Matched Database Name",
        "Category",
        "Confidence",
        "Rationale",
        "Url",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(4, i, h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = HEADER_BORDER

    for r, m in enumerate(matches, start=5):
        ws.cell(r, 1, m.ifc_material_name)
        if m.okobaudat:
            ws.cell(r, 2, m.okobaudat.name)
            ws.cell(r, 3, m.okobaudat.category)
            url_cell = ws.cell(r, 6, m.okobaudat.url)
            url_cell.hyperlink = m.okobaudat.url
            url_cell.font = LINK_FONT
        else:
            ws.cell(r, 2, "— no match —").font = Font(italic=True, color="808080")
        ws.cell(r, 4, m.confidence or "")
        ws.cell(r, 5, m.rationale or "")

        fill = _CONFIDENCE_FILL.get((m.confidence or "").lower())
        if fill:
            ws.cell(r, 4).fill = fill
        ws.row_dimensions[r].height = 25

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 40


def _format_context(context: dict | None) -> str:
    if not context:
        return ""
    parts = []
    for k, v in context.items():
        if isinstance(v, list):
            v = ", ".join(str(x) for x in v[:5]) + (" …" if len(v) > 5 else "")
        parts.append(f"{k}: {v}")
    return " | ".join(parts)


# --- Materials & Building Elements ------------------------------------------

_LOGO_PATH = Path(__file__).resolve().parent.parent / "assets" / "logo.png"


def _white_header_band(ws: Worksheet, last_col: int) -> None:
    """Fill rows 1-9 with white so gridlines vanish behind the logo / project
    info block. Gridlines stay on for the data tables below."""
    for r in range(1, 10):
        for c in range(1, last_col + 1):
            ws.cell(r, c).fill = WHITE_FILL


def _project_header(ws: Worksheet, project: Project) -> None:
    # Logo sits in cols B-C. Project info goes in col F+ so the two don't
    # overlap regardless of auto-width on data columns below.
    if _LOGO_PATH.exists():
        try:
            img = Image(str(_LOGO_PATH))
            img.width, img.height = 400, 72
            ws.add_image(img, "B3")
        except Exception as e:
            print(f"logo embed failed: {e}", file=sys.stderr)

    ws.cell(6, 2, "Material Take-Off").font = TITLE_FONT
    ws.cell(7, 2, "Do you have feedback or need help? Not what you expected?")
    link_cell = ws.cell(8, 2, "Write us at hello@nio.energy!")
    link_cell.hyperlink = "mailto:hello@nio.energy"
    link_cell.font = LINK_FONT

    ws.cell(3, 6, project.organisation or "")
    ws.cell(5, 6, project.name or "").font = Font(bold=True, size=16)
    ws.cell(6, 6, project.description or "").font = TITLE_FONT
    ws.cell(8, 6, " | ".join(project.buildings)).font = TITLE_FONT

    # Reserve room for the logo + give the project-info column some width.
    for col_letter, width in (
        ("A", 3),
        ("B", 18),
        ("C", 18),
        ("D", 18),
        ("E", 4),
        ("F", 35),
    ):
        if (
            ws.column_dimensions[col_letter].width is None
            or ws.column_dimensions[col_letter].width < width
        ):
            ws.column_dimensions[col_letter].width = width


def _section_banner(ws: Worksheet, label: str, headers: list[str]) -> None:
    ws.cell(10, 1, label)
    last_col = len(headers)

    for col in range(1, last_col + 1):
        for row in (10, 11):
            cell = ws.cell(row, col)
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = CENTER if row == 10 else MIDDLE
            cell.border = HEADER_BORDER

    # Banner row spans the full column width (matches the target xlsx).
    ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=last_col)

    for i, header in enumerate(headers, start=1):
        ws.cell(11, i, header)


def _write_material_groups(ws: Worksheet, groups: list[MaterialGroup]) -> None:
    row = 12
    for group in groups:
        ws.cell(row, 1, group.name)
        if group.okobaudat:
            ws.cell(row, 2, group.okobaudat.name)
            ws.cell(row, 3, group.okobaudat.category)
            ws.cell(row, 10, group.okobaudat.conformity)
            url_cell = ws.cell(row, 11, group.okobaudat.url)
            url_cell.hyperlink = group.okobaudat.url
            url_cell.font = LINK_FONT
        else:
            ws.cell(row, 3, group.category)

        row = _fill_quantities_for_group(ws, group, row) + 1


def _write_elements(ws: Worksheet, elements: list[Element]) -> None:
    row = 12
    for element in elements:
        if not element.materials:
            continue
        ws.cell(row, 1, element.guid)
        ws.cell(row, 2, element.name)
        ws.cell(row, 3, element.type)

        for material in element.materials:
            ws.cell(row, 4, material.name)
            ws.cell(row, 5, material.parent_name)
            if material.okobaudat:
                ws.cell(row, 6, material.okobaudat.name)
                ws.cell(row, 7, material.okobaudat.category)
            else:
                ws.cell(row, 7, material.category)

            row = _fill_quantities_for_material(ws, material, row) + 1


def _fill_quantities_for_group(ws: Worksheet, group: MaterialGroup, row: int) -> int:
    if group.okobaudat is None or group.aggregated_quantities is None:
        return row
    basis = group.okobaudat.calculation_basis
    return _dispatch_fill(
        ws,
        row,
        basis,
        group.aggregated_quantities,
        weights_lookup=_weights_index(group.aggregated_quantities),
        lca_lookup=_lca_index(group.aggregated_lca),
    )


def _fill_quantities_for_material(ws: Worksheet, material: Material, row: int) -> int:
    if material.okobaudat is None or material.quantities is None:
        return row
    basis = material.okobaudat.calculation_basis
    return _dispatch_fill(
        ws,
        row,
        basis,
        material.quantities,
        weights_lookup=_weights_index(material.quantities),
        lca_lookup=_lca_index(material.lca),
    )


def _dispatch_fill(
    ws: Worksheet,
    row: int,
    basis: str,
    q: Quantities,
    weights_lookup: dict[str, float],
    lca_lookup: dict[str, tuple[float | None, float | None]],
) -> int:
    if basis == "Volume":
        for v in q.volumes:
            if v.name is None:
                continue
            row = _emit_quantity_row(
                ws,
                row,
                v.name,
                v.value,
                v.unit or "",
                weights_lookup.get(v.name),
                lca_lookup.get(v.name),
            )
    elif basis == "Area":
        for a in q.areas:
            if a.name is None:
                continue
            row = _emit_quantity_row(
                ws,
                row,
                a.name,
                a.value,
                a.unit or "",
                weights_lookup.get(a.name),
                lca_lookup.get(a.name),
            )
    elif basis == "Weight":
        volume_by_name = {v.name: v for v in q.volumes if v.name}
        for w in q.weights:
            if w.name is None:
                continue
            vol = volume_by_name.get(w.name)
            row = _emit_quantity_row(
                ws,
                row,
                w.name,
                vol.value if vol else None,
                vol.unit if vol else "",
                w.value,
                lca_lookup.get(w.name),
            )
    elif basis == "Length" and q.lengths is not None:
        row = _emit_quantity_row(
            ws,
            row,
            "Length",
            q.lengths.length,
            q.lengths.unit,
            None,
            lca_lookup.get("Length"),
        )
    # NumberOfPieces intentionally not rendered — matches the C# behaviour.
    return row


def _emit_quantity_row(
    ws: Worksheet,
    row: int,
    quantity_name: str,
    value: float | None,
    unit: str,
    weight: float | None,
    lca: tuple[float | None, float | None] | None,
) -> int:
    col = QUANTITY_START_COL[ws.title]
    row += 1
    ws.cell(row, col, quantity_name)
    _num(ws.cell(row, col + 1, value))
    ws.cell(row, col + 2, unit)
    _num(ws.cell(row, col + 3, weight))
    embodied, whole_life = lca if lca else (None, None)
    _num(ws.cell(row, col + 4, embodied))
    _num(ws.cell(row, col + 5, whole_life))
    return row


def _num(cell) -> None:
    if cell.value is not None:
        cell.number_format = NUMBER_FORMAT


def _weights_index(q: Quantities) -> dict[str, float]:
    return {w.name: w.value for w in q.weights if w.name}


def _lca_index(lca) -> dict[str, tuple[float | None, float | None]]:
    if not lca:
        return {}
    return {
        entry.reference_quantity: (entry.embodied_carbon, entry.whole_life_carbon)
        for entry in lca
        if entry.reference_quantity
    }


def _finalize(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        width = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=12,
        )
        ws.column_dimensions[col_letter].width = min(max(width + 2, 12), 60)


if __name__ == "__main__":
    sys.exit(main())
